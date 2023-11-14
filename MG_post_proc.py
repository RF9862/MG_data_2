import re
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

def post_processing(Main_prop, Attatch_info, save_path, filename):
    '''
    This function makes xlsx files from Boxes and Texts.
    1. Loop Boxes and Texts according to document.
    2. In a document, loop boxes and texts according to page.
    3. SR number modification
    4. Consider cell swrap, thin
    5. Save
    '''
    wb = Workbook()
    thin_border = Border(left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))    
    ## Main properties 
    ws = wb.active
    ws.title = "Main Properties"   
    col_title = list(Main_prop[0][0].keys())  
    pre_rows = 0 # considering multi tables in a page.
    for k, col in enumerate(col_title):
        ws.cell(row=pre_rows+1, column=1+k).value = col
        ws.cell(row=pre_rows+1, column=1+k).font = Font(bold=True, size="20")
        ws.cell(row=pre_rows+1, column=1+k).border = thin_border     
            
    pre_rows += 1
    for results in Main_prop:
        for result in results:
            for k, col in enumerate(col_title):
                try: ws.cell(row=pre_rows+1, column=1+k).value = result[col]
                except: pass
                
            pre_rows += 1
        
    # cell swrap, thin
    row_no = 1
    for i in ws.rows:
        for j in range(len(i)):
            ws[get_column_letter(j+1)+str(row_no)].alignment = Alignment(wrap_text=True, vertical='center',horizontal='center')
            ws.cell(row=row_no, column=j + 1).border = thin_border
        row_no = row_no + 1  

    # column width
    column_width = [30] * len(col_title)
    for i in range(len(col_title)):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width[i]
    ws.sheet_view.zoomScale = 85    

    # attatchment properties
    attatch_sheet = 'Attatchment'
    wb.create_sheet(attatch_sheet)     
    ws = wb[attatch_sheet]
    col_title = ['CIN', 'Shareholder Name', 'Shareholder Address', 'No of Shares Held_history_ik', 'File name']
          
    for i in range(len(col_title)):      
        ws.cell(1,i+1).value = col_title[i]
        ws.cell(1,i+1).font = Font(bold=True)
    pre_rows = 2 
    col_cnt = 5
    for k in range(len(Attatch_info)):
        attatch_info = Attatch_info[k]
        file_name = filename[k]
        share_name, addr, no_share = attatch_info[0], attatch_info[1], attatch_info[2]
        len_share = len(share_name)
        if len_share == 0:
            print("Error, Not share name")
        else:
            for i in range(len_share):
                try: ws.cell(pre_rows+i,1).value = Main_prop[k][0]['Cin']
                except: pass
                ws.cell(pre_rows+i,2).value = share_name[i]
                try: ws.cell(pre_rows+i,3).value = no_share[i]
                except: pass
                try: ws.cell(pre_rows+i,4).value = addr[i]
                except: pass
                ws.cell(pre_rows+i,5).value = file_name
                # ws.cell(pre_rows+i,8).value = k
        pre_rows = ws.max_row + 1

    for i in range(ws.max_row):
        # ws = heading_insert(ws, pre_rows+i, heading)     
        for j in range(col_cnt):
            ws.cell(row=1+i, column=1+j).border = thin_border
    # cell swrap, thin
    row_no = 1
    for i in ws.rows:
        for j in range(len(i)):
            ws[get_column_letter(j+1)+str(row_no)].alignment = Alignment(wrap_text=True, vertical='center',horizontal='center')
        row_no = row_no + 1  

    # column width
    column_width = [30, 30, 20, 20, 40]
    for i in range(col_cnt):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width[i]
    ws.sheet_view.zoomScale = 85
    wb.save(save_path)